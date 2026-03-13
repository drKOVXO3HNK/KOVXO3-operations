from datetime import date
from io import BytesIO
from fastapi import FastAPI, Depends, Form, Request, HTTPException, UploadFile, File, Header
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, select
from openpyxl import Workbook, load_workbook

from .db import init_db, get_session
from .models import FieldItem, CropItem, OperationItem, UserItem, AuditLogItem

app = FastAPI(title="KOVXO3 Operations")
templates = Jinja2Templates(directory="app/templates")


def get_current_user(request: Request, session: Session) -> UserItem:
    username = request.cookies.get("user")
    if not username:
        raise HTTPException(status_code=401)
    user = session.exec(select(UserItem).where(UserItem.username == username)).first()
    if not user:
        raise HTTPException(status_code=401)
    return user


def can_edit(role: str) -> bool:
    return role in {"manager", "agronom"}


def get_user_by_token(session: Session, token: str | None):
    if not token:
        return None
    return session.exec(select(UserItem).where(UserItem.api_token == token)).first()


def audit(session: Session, username: str, action: str, payload: str = ""):
    session.add(AuditLogItem(username=username, action=action, payload=payload[:4000]))
    session.commit()


@app.on_event("startup")
def on_startup():
    init_db()


@app.get("/")
def root():
    return RedirectResponse(url="/dashboard")


@app.get("/login")
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), session: Session = Depends(get_session)):
    user = session.exec(select(UserItem).where(UserItem.username == username, UserItem.password == password)).first()
    if not user:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Неверный логин/пароль"})
    resp = RedirectResponse(url="/dashboard", status_code=303)
    resp.set_cookie("user", username, httponly=True)
    return resp


@app.get("/logout")
def logout():
    resp = RedirectResponse(url="/login", status_code=303)
    resp.delete_cookie("user")
    return resp


@app.get("/dashboard")
def dashboard(request: Request, season: int | None = None, session: Session = Depends(get_session)):
    try:
        user = get_current_user(request, session)
    except Exception:
        return RedirectResponse(url="/login", status_code=303)

    season = season or date.today().year
    fields = session.exec(select(FieldItem)).all()
    crops = session.exec(select(CropItem)).all()
    ops = session.exec(select(OperationItem).where(OperationItem.season == season)).all()

    by_crop: dict[str, float] = {}
    for op in ops:
        c = next((x for x in crops if x.id == op.crop_id), None)
        key = c.name if c else "Не указана"
        by_crop[key] = by_crop.get(key, 0) + op.planned_area_ha

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "season": season,
            "fields_count": len(fields),
            "ops_count": len(ops),
            "planned_sum": round(sum(x.planned_area_ha for x in ops), 2),
            "completed_sum": round(sum(x.completed_area_ha for x in ops), 2),
            "by_crop": by_crop,
            "ops": ops,
            "fields": {f.id: f for f in fields},
            "crops": {c.id: c for c in crops},
            "user": user.username,
            "role": user.role,
            "can_edit": can_edit(user.role),
        },
    )


@app.post("/seed")
def seed_data(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    if user.role != "manager":
        raise HTTPException(status_code=403, detail="Только manager")

    if not session.exec(select(FieldItem)).first():
        session.add_all([
            FieldItem(name="Поле-1", group_name="Север", area_ha=210.5),
            FieldItem(name="Поле-2", group_name="Север", area_ha=148.7),
            FieldItem(name="Поле-3", group_name="Юг", area_ha=99.5),
        ])
    if not session.exec(select(CropItem)).first():
        session.add_all([CropItem(name="Пшеница яровая"), CropItem(name="Горох"), CropItem(name="Рапс яровой")])
    if not session.exec(select(UserItem)).first():
        session.add_all([
            UserItem(username="oleg", password="oleg123", role="manager"),
            UserItem(username="agronom1", password="agro123", role="agronom"),
            UserItem(username="operator1", password="op123", role="operator"),
        ])
    session.commit()
    audit(session, user.username, "seed", "seed demo data")
    return RedirectResponse(url="/dashboard", status_code=303)


@app.post("/operations")
def add_operation(
    request: Request,
    season: int = Form(...),
    operation_type: str = Form(...),
    field_id: int = Form(...),
    crop_id: int = Form(...),
    planned_area_ha: float = Form(...),
    status: str = Form("planned"),
    session: Session = Depends(get_session),
):
    user = get_current_user(request, session)
    if not can_edit(user.role):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    session.add(
        OperationItem(
            season=season,
            operation_type=operation_type,
            field_id=field_id,
            crop_id=crop_id,
            planned_area_ha=planned_area_ha,
            completed_area_ha=0,
            status=status,
            planned_date=date.today(),
        )
    )
    session.commit()
    audit(session, user.username, "add_operation", f"season={season},field={field_id},crop={crop_id},area={planned_area_ha}")
    return RedirectResponse(url=f"/dashboard?season={season}", status_code=303)


@app.post("/operations/{op_id}/update")
def update_operation(
    op_id: int,
    request: Request,
    season: int = Form(...),
    operation_type: str = Form(...),
    planned_area_ha: float = Form(...),
    status: str = Form(...),
    session: Session = Depends(get_session),
):
    user = get_current_user(request, session)
    if not can_edit(user.role):
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    op = session.get(OperationItem, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="Operation not found")
    op.season = season
    op.operation_type = operation_type
    op.planned_area_ha = planned_area_ha
    op.status = status
    session.add(op)
    session.commit()
    audit(session, user.username, "update_operation", f"id={op_id},status={status}")
    return RedirectResponse(url=f"/dashboard?season={season}", status_code=303)


@app.post("/operations/{op_id}/close")
def close_operation(
    op_id: int,
    request: Request,
    completed_area_ha: float = Form(...),
    completed_date: str = Form(...),
    session: Session = Depends(get_session),
):
    user = get_current_user(request, session)
    if not can_edit(user.role):
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    op = session.get(OperationItem, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="Operation not found")
    op.completed_area_ha = completed_area_ha
    op.completed_date = date.fromisoformat(completed_date)
    op.status = "done"
    session.add(op)
    session.commit()
    audit(session, user.username, "close_operation", f"id={op_id},completed={completed_area_ha}")
    return RedirectResponse(url=f"/dashboard?season={op.season}", status_code=303)


@app.post("/operations/{op_id}/delete")
def delete_operation(op_id: int, request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    if user.role != "manager":
        raise HTTPException(status_code=403, detail="Только manager")
    op = session.get(OperationItem, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="Operation not found")
    season = op.season
    session.delete(op)
    session.commit()
    audit(session, user.username, "delete_operation", f"id={op_id}")
    return RedirectResponse(url=f"/dashboard?season={season}", status_code=303)


@app.post("/import/operations")
def import_operations(request: Request, season: int = Form(...), file: UploadFile = File(...), session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    if user.role != "manager":
        raise HTTPException(status_code=403, detail="Только manager")

    wb = load_workbook(filename=BytesIO(file.file.read()), data_only=True)
    ws = wb.active

    fields = {f.name: f.id for f in session.exec(select(FieldItem)).all()}
    crops = {c.name: c.id for c in session.exec(select(CropItem)).all()}

    created = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        operation_type, field_name, crop_name, planned_area = row[0], row[1], row[2], row[3]
        if field_name not in fields or crop_name not in crops:
            continue
        session.add(OperationItem(
            season=season,
            operation_type=str(operation_type),
            field_id=fields[field_name],
            crop_id=crops[crop_name],
            planned_area_ha=float(planned_area or 0),
            completed_area_ha=0,
            status="planned",
            planned_date=date.today(),
        ))
        created += 1
    session.commit()
    audit(session, user.username, "import_operations", f"season={season},created={created},file={file.filename}")
    return RedirectResponse(url=f"/dashboard?season={season}", status_code=303)


@app.get("/fields/{field_id}")
def field_card(field_id: int, request: Request, season: int | None = None, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    season = season or date.today().year
    field = session.get(FieldItem, field_id)
    if not field:
        return RedirectResponse(url=f"/dashboard?season={season}", status_code=303)
    ops = session.exec(select(OperationItem).where(OperationItem.field_id == field_id, OperationItem.season == season)).all()
    crops = {c.id: c for c in session.exec(select(CropItem)).all()}
    return templates.TemplateResponse("field.html", {
        "request": request,
        "season": season,
        "field": field,
        "ops": ops,
        "crops": crops,
        "user": user.username,
        "role": user.role,
    })


@app.get("/audit")
def audit_page(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    if user.role != "manager":
        raise HTTPException(status_code=403, detail="Только manager")
    logs = session.exec(select(AuditLogItem).order_by(AuditLogItem.id.desc())).all()
    return templates.TemplateResponse("audit.html", {"request": request, "logs": logs, "user": user.username, "role": user.role})


@app.get("/export/operations.xlsx")
def export_operations(request: Request, season: int, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)
    fields = {f.id: f.name for f in session.exec(select(FieldItem)).all()}
    crops = {c.id: c.name for c in session.exec(select(CropItem)).all()}
    ops = session.exec(select(OperationItem).where(OperationItem.season == season)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Season {season}"
    ws.append(["ID", "Season", "Operation", "Field", "Crop", "Planned ha", "Completed ha", "Status", "Planned date", "Completed date"])
    for o in ops:
        ws.append([
            o.id, o.season, o.operation_type, fields.get(o.field_id, o.field_id), crops.get(o.crop_id, "Не указана"),
            o.planned_area_ha, o.completed_area_ha, o.status,
            o.planned_date.isoformat() if o.planned_date else "",
            o.completed_date.isoformat() if o.completed_date else "",
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=operations_{season}.xlsx"},
    )


@app.get("/api/token/me")
def api_token_me(request: Request, session: Session = Depends(get_session)):
    user = get_current_user(request, session)
    return {"username": user.username, "role": user.role, "api_token": user.api_token}


@app.post("/api/token/rotate")
def api_token_rotate(request: Request, session: Session = Depends(get_session)):
    import secrets
    user = get_current_user(request, session)
    user.api_token = secrets.token_hex(16)
    session.add(user)
    session.commit()
    audit(session, user.username, "rotate_token", "")
    return {"api_token": user.api_token}


@app.post("/api/operations")
def api_add_operation(
    season: int,
    operation_type: str,
    field_id: int,
    crop_id: int,
    planned_area_ha: float,
    x_api_token: str | None = Header(default=None),
    session: Session = Depends(get_session),
):
    user = get_user_by_token(session, x_api_token)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid API token")
    if not can_edit(user.role):
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    op = OperationItem(
        season=season,
        operation_type=operation_type,
        field_id=field_id,
        crop_id=crop_id,
        planned_area_ha=planned_area_ha,
        completed_area_ha=0,
        status="planned",
        planned_date=date.today(),
    )
    session.add(op)
    session.commit()
    session.refresh(op)
    audit(session, user.username, "api_add_operation", f"id={op.id}")
    return {"id": op.id, "status": op.status}


@app.get("/api/report/{season}")
def report(request: Request, season: int, session: Session = Depends(get_session)):
    _ = get_current_user(request, session)
    ops = session.exec(select(OperationItem).where(OperationItem.season == season)).all()
    return {
        "season": season,
        "operations": len(ops),
        "planned_area_ha": round(sum(x.planned_area_ha for x in ops), 2),
        "completed_area_ha": round(sum(x.completed_area_ha for x in ops), 2),
    }
